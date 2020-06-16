#coding=gbk
'''
Created on Dec 3, 2018

@author: ben
'''
from skimage import io,transform
import glob
import os
import tensorflow as tf
import numpy as np
import time
import xlrd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
import cv2
#import xlwt
import os


samples_answer_excel =  'E:/AI/ctc/预测模型分类.xlsx'
book_answer_samples = xlrd.open_workbook(samples_answer_excel)#打开一个excel
sheet_answer_samples = book_answer_samples.sheet_by_index(0)#根据顺序获取sheet这个是读取第一个sheet

samples_name_excel =  'E:/AI/ctc/预测模型分类.xlsx'
book_name_samples = xlrd.open_workbook(samples_name_excel)#打开一个excel
sheet_name_samples = book_name_samples.sheet_by_index(1)#根据顺序获取sheet这个是读取第一个sheet

cep8_weight = 0.9
red_weight = 0.2
#数据集地址
path='E:/AI/ctc/CTC/'

test_path='E:/AI/ctc/test_ctc/'
#模型保存地址
model_path='./log/model.ckpt'



#将所有的图片resize成100*100
w=300  #2728
h=300  #2192
c=3
batch_size=4

def read_one_image(path):
    img = io.imread(path)
    img = transform.resize(img,(w,h))
    return np.asarray(img)

def read_test_img(sample_name,data):
    count1 = 0
    for row_name_samples in sheet_name_samples.get_rows():
        if str(row_name_samples[12].value) == "test":
            sample_name_valid = row_name_samples[0].value
            folder = str(row_name_samples[0].value)   
            '''         
            answer = -1
            for row_samples in sheet_answer_samples.get_rows():
                if folder == str(row_samples[1].value):
                    answer = row_samples[8].value
                    if int(answer) >= 2:
                        answer = 1
                    else:
                        answer = 0
                    break
            '''
            all_path = path + folder + "\\cep8.jpg"
            all_path_dapi = path + folder + "\\dapi.jpg"
            all_path_red = path + folder + "\\red.jpg"
            if os.path.exists(all_path) and os.path.exists(all_path_dapi) and  os.path.exists(all_path_red) :
                #print('%s'%(folder), " ", str(answer))
                
                count1 = count1 + 1
                
                if count1 <50 or count1>150:
                   
                    continue
                
                print(str(folder)," ", str(count1))
                #file.write('%s'%(folder) +" " + str(answer) + "\n")
                #ws.cell(row=row_excel, column=1).value = folder
                #ws.cell(row=row_excel, column=2).value = answer

                #data1 = read_one_image(all_path)
                img = io.imread(all_path)
                img_dapi = io.imread(all_path_dapi)
                img_red = io.imread(all_path_red)
                #img = cv2.addWeighted(img, cep8_weight, img_dapi, 1.0 - cep8_weight, 0) 
                #img = cv2.addWeighted(img, 1.0 - red_weight, img_red, red_weight, 0) 
                
                
                #cv2.namedWindow('img_mix',0)
                #cv2.imshow('img_mix', img)
                
               
                img = transform.resize(img,(w,h))
                
                img = img * 255   #将图片的取值范围改成（0~255）
                img = img.astype(np.uint8) 

                full_path_write = path + str(folder) + '/img_mix1.jpeg'  
                print(" this is ",full_path_write)
                #cv2.imwrite(full_path_write,img)
                cv2.imencode('.jpg',img)[1].tofile(full_path_write)
                data1 = np.asarray(img)
                sample_name.append(folder)
                data.append(data1)

#读取图片
def read_img(path):
    file = open('log.txt', 'w')
    #file1 = xlwt.Workbook()
    excel_file=load_workbook(filename=r'E:\AI\ctc\log_template.xlsx')
    ws = excel_file.get_sheet_by_name("Sheet2")
    row_excel = 1
    imgs=[]
    labels=[]    
    for row_name_samples in sheet_name_samples.get_rows():
        if str(row_name_samples[12].value) == "train":
            folder = str(row_name_samples[0].value)        
            answer = -1                
            if int(row_name_samples[9].value) >= 2:
                answer = 1
            else:
                answer = 0
    
            all_path = path + folder + "\\cep8.jpg"
            all_path_dapi = path + folder + "\\dapi.jpg"
            all_path_red = path + folder + "\\red.jpg"
            if answer != -1 and os.path.exists(all_path) and os.path.exists(all_path_dapi) and  os.path.exists(all_path_red):
                #im = all_path
                #im_dapi = all_path_dapi
                #im_red = all_path_red
                print('reading the images:%s'%(all_path)," answer: ", str(answer))
                file.write('reading the images:%s'%(all_path)+" answer: "+ str(answer))
                ws.cell(row=row_excel, column=1).value = folder
                ws.cell(row=row_excel, column=2).value = str(answer)
                row_excel += 1
                
                img=io.imread(all_path)
                img_dapi = io.imread(all_path_dapi)
                img_red = io.imread(all_path_red)
                #img = cv2.addWeighted(img, cep8_weight, img_dapi, 1.0 - cep8_weight, 0)
                #img = cv2.addWeighted(img, 1.0 - red_weight, img_red, red_weight, 0) 
                img=transform.resize(img,(w,h))
                imgs.append(img)
                labels.append(answer)              
    file.close()
    excel_file.save(filename= 'E:\AI\ctc\log.xlsx')                
    '''
    cate=[x for x in os.listdir(path) if os.path.isdir(path+x)]
    imgs=[]
    labels=[]
    for idx,folder in enumerate(cate):
        for im in glob.glob(path + folder+'/*.jpg'):
            print('reading the images:%s'%(im))
            img=io.imread(im)
            img=transform.resize(img,(w,h))
            imgs.append(img)

            labels.append(answer)
            
     '''       
    return np.asarray(imgs,np.float32),np.asarray(labels,np.int32)


def inference(input_tensor, train, regularizer):
    with tf.variable_scope('layer1-conv1'):
        conv1_weights = tf.get_variable("weight",[5,5,3,32],initializer=tf.truncated_normal_initializer(stddev=0.1))
        conv1_biases = tf.get_variable("bias", [32], initializer=tf.constant_initializer(0.0))
        conv1 = tf.nn.conv2d(input_tensor, conv1_weights, strides=[1, 1, 1, 1], padding='SAME')
        relu1 = tf.nn.relu(tf.nn.bias_add(conv1, conv1_biases))

    with tf.name_scope("layer2-pool1"):
        pool1 = tf.nn.max_pool(relu1, ksize = [1,2,2,1],strides=[1,2,2,1],padding="VALID")

    with tf.variable_scope("layer3-conv2"):
        conv2_weights = tf.get_variable("weight",[5,5,32,64],initializer=tf.truncated_normal_initializer(stddev=0.1))
        conv2_biases = tf.get_variable("bias", [64], initializer=tf.constant_initializer(0.0))
        conv2 = tf.nn.conv2d(pool1, conv2_weights, strides=[1, 1, 1, 1], padding='SAME')
        relu2 = tf.nn.relu(tf.nn.bias_add(conv2, conv2_biases))

    with tf.name_scope("layer4-pool2"):
        pool2 = tf.nn.max_pool(relu2, ksize=[1, 2, 2, 1], strides=[1, 2, 2, 1], padding='VALID')

    with tf.variable_scope("layer5-conv3"):
        conv3_weights = tf.get_variable("weight",[3,3,64,128],initializer=tf.truncated_normal_initializer(stddev=0.1))
        conv3_biases = tf.get_variable("bias", [128], initializer=tf.constant_initializer(0.0))
        conv3 = tf.nn.conv2d(pool2, conv3_weights, strides=[1, 1, 1, 1], padding='SAME')
        relu3 = tf.nn.relu(tf.nn.bias_add(conv3, conv3_biases))

    with tf.name_scope("layer6-pool3"):
        pool3 = tf.nn.max_pool(relu3, ksize=[1, 2, 2, 1], strides=[1, 2, 2, 1], padding='VALID')

    with tf.variable_scope("layer7-conv4"):
        conv4_weights = tf.get_variable("weight",[3,3,128,128],initializer=tf.truncated_normal_initializer(stddev=0.1))
        conv4_biases = tf.get_variable("bias", [128], initializer=tf.constant_initializer(0.0))
        conv4 = tf.nn.conv2d(pool3, conv4_weights, strides=[1, 1, 1, 1], padding='SAME')
        relu4 = tf.nn.relu(tf.nn.bias_add(conv4, conv4_biases))

    with tf.name_scope("layer8-pool4"):
        pool4 = tf.nn.max_pool(relu4, ksize=[1, 2, 2, 1], strides=[1, 2, 2, 1], padding='VALID')
        nodes = 6*6*128 * 9
        reshaped = tf.reshape(pool4,[-1,nodes])

    with tf.variable_scope('layer9-fc1'):
        fc1_weights = tf.get_variable("weight", [nodes, 1024],
                                      initializer=tf.truncated_normal_initializer(stddev=0.1))
        if regularizer != None: tf.add_to_collection('losses', regularizer(fc1_weights))
        fc1_biases = tf.get_variable("bias", [1024], initializer=tf.constant_initializer(0.1))

        fc1 = tf.nn.relu(tf.matmul(reshaped, fc1_weights) + fc1_biases)
        if train: fc1 = tf.nn.dropout(fc1, 0.5)

    with tf.variable_scope('layer10-fc2'):
        fc2_weights = tf.get_variable("weight", [1024, 512],
                                      initializer=tf.truncated_normal_initializer(stddev=0.1))
        if regularizer != None: tf.add_to_collection('losses', regularizer(fc2_weights))
        fc2_biases = tf.get_variable("bias", [512], initializer=tf.constant_initializer(0.1))

        fc2 = tf.nn.relu(tf.matmul(fc1, fc2_weights) + fc2_biases)
        if train: fc2 = tf.nn.dropout(fc2, 0.5)

    with tf.variable_scope('layer11-fc3'):
        fc3_weights = tf.get_variable("weight", [512, 2],
                                      initializer=tf.truncated_normal_initializer(stddev=0.1))
        if regularizer != None: tf.add_to_collection('losses', regularizer(fc3_weights))
        fc3_biases = tf.get_variable("bias", [2], initializer=tf.constant_initializer(0.1))
        logit = tf.matmul(fc2, fc3_weights) + fc3_biases

    return logit



#定义一个函数，按批次取数据
def minibatches(inputs=None, targets=None, batch_size=None, shuffle=False):
    assert len(inputs) == len(targets)
    if shuffle:
        indices = np.arange(len(inputs))
        np.random.shuffle(indices)
    for start_idx in range(0, len(inputs) - batch_size + 1, batch_size):
        if shuffle:
            excerpt = indices[start_idx:start_idx + batch_size]
        else:
            excerpt = slice(start_idx, start_idx + batch_size)
        yield inputs[excerpt], targets[excerpt]









def trainning_process():
    
    data,label=read_img(path)
    #打乱顺序
    num_example=data.shape[0]
    arr=np.arange(num_example)
    np.random.shuffle(arr)
    data=data[arr]
    label=label[arr]
    
    #将所有数据分为训练集和验证集
    ratio=0.8
    s=np.int(num_example*ratio)
    x_train=data[:s]
    y_train=label[:s]
    x_val=data[s:]
    y_val=label[s:]
    
    #ben
    x_train=data
    y_train=label
    #ben

    
    #-----------------构建网络----------------------
    #占位符
    x=tf.placeholder(tf.float32,shape=[None,w,h,c],name='x')
    y_=tf.placeholder(tf.int32,shape=[None,],name='y_')
    
    
    regularizer = tf.contrib.layers.l2_regularizer(0.0001)
    logits = inference(x,False,regularizer)
    #logits = tf.reshape(logits, [-1])
    #---------------------------网络结束---------------------------
    
    #(小处理)将logits乘以1赋值给logits_eval，定义name，方便在后续调用模型时通过tensor名字调用输出tensor
    b = tf.constant(value=1,dtype=tf.float32)
    logits_eval = tf.multiply(logits,b,name='logits_eval') 
    
    loss=tf.nn.sparse_softmax_cross_entropy_with_logits(logits=logits, labels=y_)
    train_op=tf.train.AdamOptimizer(learning_rate=0.001).minimize(loss)
    correct_prediction = tf.equal(tf.cast(tf.argmax(logits,1),tf.int32), y_)    
    acc= tf.reduce_mean(tf.cast(correct_prediction, tf.float32))
    
    
    
    
    
    #训练和测试数据，可将n_epoch设置更大一些
    
    n_epoch=100                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                           
    
    saver=tf.train.Saver()
    sess=tf.Session()  
    sess.run(tf.global_variables_initializer())
    train_num=0
    for epoch in range(n_epoch):
        start_time = time.time()
        train_num = train_num + 1
        #training
        train_loss, train_acc, n_batch = 0, 0, 0
        for x_train_a, y_train_a in minibatches(x_train, y_train, batch_size, shuffle=True):
            _,err,ac=sess.run([train_op,loss,acc], feed_dict={x: x_train_a, y_: y_train_a})
            train_loss += err; train_acc += ac; n_batch += 1
        print("   train loss: %f" % (np.sum(train_loss)/ n_batch))
        print("   train acc: %f" % (np.sum(train_acc)/ n_batch))
        print(" train num is %d" % train_num)
        '''
        #validation
        val_loss, val_acc, n_batch = 0, 0, 0
        for x_val_a, y_val_a in minibatches(x_val, y_val, batch_size, shuffle=False):
            err, ac = sess.run([loss,acc], feed_dict={x: x_val_a, y_: y_val_a})
            val_loss += err; val_acc += ac; n_batch += 1
        print("   validation loss: %f" % (np.sum(val_loss)/ n_batch))
        print("   validation acc: %f" % (np.sum(val_acc)/ n_batch))
        '''
    print("this is the end")
    saver.save(sess,model_path)
    sess.close()




#########prediction
def prediction():
    file = open('log.txt', 'w')
    excel_file=load_workbook(filename=r'E:\AI\ctc\预测模型分类.xlsx')
    ws = excel_file.get_sheet_by_name("Data")
    
    #path1 = "E:/temp/111.5.白玉.1.cep8 (1).jpg"
    path2 = "E:/temp/111.5.白玉.1.cep8 (2).jpg"
    path3 = "E:/temp/654.1.白玉.2.cep8.jpg"
    path4 = "E:/temp/111.5.白玉.1.cep8 (4).jpg"
    path5 = "E:/temp/543.1.白洁峰.1.cep8.jpg"
    
    test_sample_name = "魏勇"
    
    all_path_test = path + test_sample_name + "\\cep8.jpg"
    path1 = all_path_test
    #test_data,test_label=read_img(test_path)
    
    flower_dict = {0:'0',1:'1'}
    
    #w=100
    #h=100
    #c=3
    

    
    with tf.Session() as sess:
        data = []
        sample_name = []

        read_test_img(sample_name,data)    
        #data = test_data
    
        print("  middle: ")
    
    
        saver = tf.train.import_meta_graph('./log/model.ckpt.meta')
        saver.restore(sess,tf.train.latest_checkpoint('./log/'))
    
        graph = tf.get_default_graph()
        x = graph.get_tensor_by_name("x:0")
        feed_dict = {x:data}
    
        logits = graph.get_tensor_by_name("logits_eval:0")
    
        classification_result = sess.run(logits,feed_dict)
    
        #打印出预测矩阵
        print(classification_result)
        #打印出预测矩阵每一行最大值的索引
        print(tf.argmax(classification_result,1).eval())
        #根据索引通过字典对应花的分类
        output = []
        output = tf.argmax(classification_result,1).eval()
        row_excel = 1
        for i in range(len(output)):
            print("第",i+1,"个: "+flower_dict[output[i]])
            for cx in range(ws.min_row ,ws.max_row):
                if (str(ws.cell(row=cx, column=1).value) == str(sample_name[i])):
                    print ("sample_name[",str(cx),"]: ",str(sample_name[i])," output = ", str(str(flower_dict[output[i]])))
                    ws.cell(row=cx, column=11).value = str(flower_dict[output[i]])


    file.close()
    print(path + 'log.xlsx')
    excel_file.save(filename='E:\AI\ctc\预测模型分类.xlsx')  
    #excel_file.save(filename=path + 'log.xlsx')   
    print ("this is the end of test")
#trainning_process()
prediction()


'''
def write_excel():
    ex=load_workbook(filename=r'E:\AI\ctc\log.xlsx')
    print('open excel success!')
    ws = ex.get_sheet_by_name("Sheet1")
    print('open sheet1 success!')
    ws.cell(row=4, column=2).value = 'hupi2222'
    print('write values success!')
    ex.save(filename='E:\AI\ctc\log.xlsx')
    print('save success!')


#write_excel()
'''