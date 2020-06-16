# -*- coding: utf-8 -*-  
'''
Created on Sep 26, 2018

@author: ben
'''
#-*- coding: UTF-8 -*- 
from bokeh.util.paths import ROOT_DIR
save_result = True
'''
Author: Steve Wang
Time: 2017/12/8 10:00
Environment: Python 3.6.2 |Anaconda 4.3.30 custom (64-bit) Opencv 3.3
'''

import cv2
import numpy as np
import os

from skimage import io,transform
import glob
import tensorflow as tf
import time
import xlrd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter


#samples_name_excel =  'E:/AI/ctc/result.xlsx'
#book_name_samples = xlrd.open_workbook(samples_name_excel)#打开一个excel
#sheet_name_samples = book_name_samples.sheet_by_index(1)#根据顺序获取sheet这个是读取第一个sheet
Result_XLSX = r'E:\AI\ctc\result.xlsx'
excel_file=load_workbook(filename=Result_XLSX)
ws = excel_file.get_sheet_by_name("Data1")

def get_image(path):
    #��ȡͼƬ
    #path_code = path.
    #img=cv2.imread(path)
    img = cv2.imdecode(np.fromfile(path,dtype=np.uint8),-1)
    if img.all() == None:
        print ("--------------------------------look here")
        return None,None
    print(" path in get_image: ", path)


    return img

def get_gray(img):
    #��ȡͼƬ
    #path_code = path.
    #img=cv2.imread(path)

    gray=cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)

    return gray

def Gaussian_Blur(gray):
    # ��˹ȥ��
    blurred = cv2.GaussianBlur(gray, (5, 5),0)

    return blurred

def Sobel_gradient(blurred):
    # ���ȶ�����������x��y�����ݶ�
    gradX = cv2.Sobel(blurred, ddepth=cv2.CV_32F, dx=1, dy=0)
    gradY = cv2.Sobel(blurred, ddepth=cv2.CV_32F, dx=0, dy=1)

    gradient = cv2.subtract(gradX, gradY)
    gradient = cv2.convertScaleAbs(gradient)

    return gradX, gradY, gradient

def Thresh_and_blur(gradient,threshold1):

    blurred = cv2.GaussianBlur(gradient, (5,5),0)
    #(_, thresh) = cv2.threshold(blurred, 90, 255, cv2.THRESH_BINARY)
    (_, thresh) = cv2.threshold(blurred, threshold1, 255, cv2.THRESH_BINARY)
    return thresh

def image_morphology(thresh):
    # ����һ����Բ�˺���
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (2, 2))
    #print(" kernel is",kernel)
    # ִ��ͼ����̬ѧ, ϸ��ֱ�Ӳ��ĵ����ܼ�
    #closed = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel) 
    closed = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, kernel) 
    
    closed = cv2.erode(closed, kernel, iterations=2)
    #closed = cv2.dilate(closed, kernel, iterations=1)
    

    return closed

def findcnts_and_box_point(closed, image_size_threshold = 0.0):
    # ����opencv3���ص�����������
    (_, cnts, _) = cv2.findContours(closed.copy(), 
        cv2.RETR_LIST, 
        cv2.CHAIN_APPROX_SIMPLE)
    #print (" cnts: ",cnts)
    sorted_list = sorted(cnts, key=cv2.contourArea, reverse=True)
    box_list = []
    if(len(sorted_list)>0):
        for list_index in sorted_list:
            c = list_index
            rect = cv2.minAreaRect(c)
            print("rect[1][0]: ",rect[1][0],"rect[1][1]: ",rect[1][1],"area: ", rect[1][0] * rect[1][1]  )
            if  rect[1][0] * rect[1][1] > image_size_threshold:
                print (" enter: ", "image_size_threshold = ", image_size_threshold)
                box_list.append(np.int0(cv2.boxPoints(rect)))
            
    #c = sorted(cnts, key=cv2.contourArea, reverse=True)[2]
    # compute the rotated bounding box of the largest contour
    #rect = cv2.minAreaRect(c)
    #box = np.int0(cv2.boxPoints(rect))

    return box_list

def drawcnts_and_cut(original_img, box_list):
    draw_img = original_img.copy()
    crop_img_list = []
    for box in box_list:
        # ��Ϊ��������м�ǿ���ƻ��ԣ�������Ҫ��img.copy()�ϻ�
        # draw a bounding box arounded the detected barcode and display the image
        draw_img = cv2.drawContours(draw_img, [box], -1, (0, 0, 255), 3)
    
        Xs = [i[0] for i in box]
        Ys = [i[1] for i in box]
        x1 = min(Xs)     
        x2 = max(Xs)
        y1 = min(Ys)
        y2 = max(Ys)
        x1 -= 2
        x2 += 2
        y1 -= 2
        y2 += 2
        hight = y2 - y1
        width = x2 - x1
        crop_img = original_img[y1:y1+hight, x1:x1+width]
        print("x1",x1,"x2",x2,"y1",y1,"y2",y2,"hight",hight,"width",width, "area: ", hight * width)
        crop_img_list.append(crop_img)

    return draw_img, crop_img_list


def image_process(img,save_path_draft,threshold1, image_size_threshold = 0.0):
    #original_img, gray = get_image(img)
    original_img =img
    gray = get_gray(img)
    if original_img.all() ==None:
        return
    blurred = Gaussian_Blur(gray)
    gradX, gradY, gradient = Sobel_gradient(blurred)
    gradient = blurred
    thresh = Thresh_and_blur(gradient,threshold1)
    closed = image_morphology(thresh)
    box_list = findcnts_and_box_point(closed, image_size_threshold)
    #for box in box_list: 
    if True:     
        draw_img, crop_img_list = drawcnts_and_cut(original_img,box_list)
    
        # ����һ�㣬�����Ƕ���ʾ��������
    
        #cv2.imshow('original_img', original_img)
        
        #cv2.namedWindow('blurred',0)
        #cv2.imshow('blurred', blurred)
        if(save_result):
            cv2.imencode('.jpg',blurred)[1].tofile(save_path_draft + "blurred.jpg")
        
        
        #cv2.namedWindow('gradX',0)
        #cv2.imshow('gradX', gradX)
        if(save_result):
            cv2.imencode('.jpg',gradX)[1].tofile(save_path_draft + "gradX.jpg")
        #cv2.namedWindow('gradY',0)
        #cv2.imshow('gradY', gradY)
        if(save_result):
            cv2.imencode('.jpg',gradY)[1].tofile(save_path_draft + "gradY.jpg")
                
        #cv2.namedWindow('final',0)
        #cv2.imshow('final', gradient)
        if(save_result):
            cv2.imencode('.jpg',gradient)[1].tofile(save_path_draft + "gradient.jpg")
         
        #cv2.namedWindow('thresh',0)
        #cv2.imshow('thresh', thresh)
        if(save_result):
            cv2.imencode('.jpg',thresh)[1].tofile(save_path_draft + "thresh.jpg")
        
        
        #cv2.namedWindow('closed',0)
        #cv2.imshow('closed', closed)
        if(save_result):
            cv2.imencode('.jpg',closed)[1].tofile(save_path_draft + "closed.jpg")
        #cv2.waitKey(100000)

    return draw_img, box_list, crop_img_list

def walk():
    ROOT_DIR_CTC = 'E:/AI/ctc/CTC/'
    save_path_small_img = ROOT_DIR_CTC
    save_path_orignal = 'E:/AI/ctc/small_image_orignal/'
    save_path_orignal_0 = 'E:/AI/ctc/0/'
    save_path_orignal_1 = 'E:/AI/ctc/1/'
    save_path_big_img = ROOT_DIR_CTC
    
    if not os.path.exists(save_path_small_img):
        os.makedirs(save_path_small_img) 

    if not os.path.exists(save_path_orignal):
        os.makedirs(save_path_orignal) 
        
    if not os.path.exists(save_path_orignal_0):
        os.makedirs(save_path_orignal_0)         
        
    if not os.path.exists(save_path_orignal_1):
        os.makedirs(save_path_orignal_1) 

    if not os.path.exists(save_path_big_img):
        os.makedirs(save_path_big_img) 

    
    filename_rgb = r'E:/AI/ctc/CTC/'
    for filename in os.listdir(filename_rgb):              #listdir
        print (str(filename))
    
    #for filename1 in glob.glob(r'E:/AI/ctc/CTC/*.exe'):
    #    print (filename1)
    
    for dirpath, dirnames, filenames in os.walk(ROOT_DIR_CTC):
        print ('Directory', dirpath, '  dirnames: ',dirnames)
        
        for filename2 in filenames:
            Last_path = dirpath.lstrip(ROOT_DIR_CTC)
            Last_path = Last_path
            #print (' File', filename2)
            if(filename2 == "dapi.jpg"):
                img_path = dirpath
                filename_dapi = "dapi"
                filename_dapi_jpg = "dapi.jpg"
                filename_cep8 = "cep8" 
                filename_cep8_jpg = "cep8.jpg"
                dapi_img_path =  dirpath +'/'+ filename_dapi_jpg
                cep8_img_path =   dirpath +'/'+ filename_cep8_jpg
                #save_path = 'E:\\AI\\ctc\\result\\result_of_'

                save_path_big_img_custom_dapi_jpg = save_path_big_img + Last_path +"/result_" + "drew_" + filename_dapi_jpg
                save_path_big_img_custom_dapi = save_path_big_img + Last_path  +"/result_" + "drew_" + filename_dapi
                ######确定文件后
                img =  get_image(dapi_img_path)
                
                try:
                    draw_img, box_list, crop_img_list = image_process(img, save_path_big_img_custom_dapi, 40, 1000)
                
                except:
                    continue
                ##########
                #cv2.namedWindow('draw_img',0)
                #cv2.imshow('draw_img', draw_img)
                #cv2.imwrite(save_path, draw_img)
                if(save_result):
                    cv2.imencode('.jpg',draw_img)[1].tofile(save_path_big_img_custom_dapi_jpg)
                '''
                number = 0
                print (" len of crop image: ",str(len(crop_img_list)))
                for crop_img in crop_img_list:
                    number = number + 1
                    crop_img_name = 'crop_img' + str(number)
                    cv2.namedWindow(crop_img_name,0)
                    cv2.imshow(crop_img_name, crop_img)
                    #
                    #cv2.imwrite(save_path, crop_img)
                '''
                #cv2.waitKey(20171219)
                try:
                    cep8_img =  get_image(cep8_img_path)
                except:
                    continue
                draw_cep8_img, crop_cep8_img_list = drawcnts_and_cut(cep8_img,box_list)
                cep8_img_num = 0
                for crop_cep8_img in crop_cep8_img_list:
                    cep8_img_num = cep8_img_num + 1
                    save_path_cep8_img = save_path_small_img + Last_path + "/result_"+ str(cep8_img_num) + "." + filename_cep8_jpg
                    save_path_cep8 = save_path_small_img + Last_path + "/result_"+ str(cep8_img_num) + "." + filename_cep8
                    print(" image_process in image:", save_path_cep8_img)
                    try:
                        draw_img1, box_list1, crop_img_list1 = image_process(crop_cep8_img, save_path_cep8, 30, 0.1)
                    except:
                        continue
                    #save drawed small image
                    if(save_result):                   
                        cv2.imencode('.jpg',draw_img1)[1].tofile(save_path_cep8_img)
                    
                    #save orignal small image
                    
                    #save_path_cep8_img_orignal = save_path_orignal + Last_path + "/"+ str(cep8_img_num) + "." + filename_cep8_jpg       
                         
                                   
                    #cv2.imencode('.jpg',crop_cep8_img)[1].tofile(save_path_cep8_img_orignal)                
                    if len(box_list1)>2:
                        file_name1 = "1_" + str(cep8_img_num) + "." + filename_cep8_jpg
                        save_path_cep8_img_orignal_1 = save_path_small_img + Last_path + "/"+ file_name1
                        cv2.imencode('.jpg',crop_cep8_img)[1].tofile(save_path_cep8_img_orignal_1)
                        row_last = ws.max_row + 1
                        ws.cell(row=row_last, column=11).value = str(save_path_small_img)
                        ws.cell(row=row_last, column=1).value = str(Last_path)
                        ws.cell(row=row_last, column=2).value = str(file_name1)
                        ws.cell(row=row_last, column=3).value = str("1")
                        #ws.cell(row=1, column=3).value = str("test")
                         
                    else:
                        file_name0 = "0_" + str(cep8_img_num) + "." + filename_cep8_jpg
                        save_path_cep8_img_orignal_0 = save_path_small_img + Last_path +"/" +file_name0
                        cv2.imencode('.jpg',crop_cep8_img)[1].tofile(save_path_cep8_img_orignal_0)
                        row_last = ws.max_row + 1
                        ws.cell(row=row_last, column=11).value = str(save_path_small_img)
                        ws.cell(row=row_last, column=1).value = str(Last_path)
                        ws.cell(row=row_last, column=2).value = str(file_name0)
                        ws.cell(row=row_last, column=3).value = str("0")
                        #ws.cell(row=2, column=3).value = str("test")
                        #excel_file.save(filename='E:\AI\ctc\result.xlsx')  
                    
                ##########
                
                
                #######
                
    excel_file.save(filename=Result_XLSX)             
    print ("this is the end")
            
walk()

