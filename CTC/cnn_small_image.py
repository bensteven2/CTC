#coding=gbk
'''
Created on Dec 3, 2018

@author: ben
'''
from skimage import io,transform
import glob
import os
import tensorflow as tf
import numpy as np
import time
import xlrd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
import cv2
import argparse
#import xlwt
import sys
sys.path.append("..")
import  AI.CNN as CNN 
import AI.CNN_model_V1_0 as CNN_model_V1_0
#Result_XLSX = r'/data2/ben/data/geneis/CTC/result.xlsx'
#excel_file=load_workbook(filename=Result_XLSX)
#ws = excel_file.get_sheet_by_name("Data1")
    
cep8_weight = 0.9
red_weight = 0.2
#数据集地址
path='/data2/ben/data/geneis/CTC/images/'

#模型保存地址
model_path='./log/model.ckpt'



#将所有的图片resize成100*100
#w=100  #2728
#h=100  #2192
c=3
batch_size=4

def read_one_image(path):
    img = io.imread(path)
    img = transform.resize(img,(w,h))
    return np.asarray(img)

def read_test_img(sample_name,data, start, end):
    Result_XLSX = r'/data2/ben/data/geneis/CTC/result.xlsx'
    book_name_samples = xlrd.open_workbook(Result_XLSX)#打开一个excel
    sheet_name_samples = book_name_samples.sheet_by_index(0)#根据顺序获取sheet这个是读取第一个sheet
    count1 = 0
    for row_name_samples in sheet_name_samples.get_rows():
        #if str(row_name_samples[4].value) != "train":
        if True:
            folder = str(row_name_samples[0].value)   
            '''         
            answer = -1
            for row_samples in sheet_answer_samples.get_rows():
                if folder == str(row_samples[1].value):
                    answer = row_samples[8].value
                    if int(answer) >= 2:
                        answer = 1
                    else:
                        answer = 0
                    break
            '''
            all_path = str(row_name_samples[10].value)  + folder + '/' + str(row_name_samples[1].value)
            if os.path.isfile(all_path):
                #print('%s'%(folder), " ", str(answer))
                
                count1 = count1 + 1
                
                if count1 < start or count1 > end:
                   
                    continue
                
                print(str(folder)," ", str(count1))
                #file.write('%s'%(folder) +" " + str(answer) + "\n")
                #ws.cell(row=row_excel, column=1).value = folder
                #ws.cell(row=row_excel, column=2).value = answer

                #data1 = read_one_image(all_path)
                img = io.imread(all_path)   
                img = transform.resize(img,(w,h))
                data1 = np.asarray(img)
                sample_name.append(folder + '/' + str(row_name_samples[1].value))
                data.append(data1)

#读取图片
def read_img(images_dir, labels_address, size_square, mode_key):
    book_name_samples = xlrd.open_workbook(labels_address)#打开一个excel
    sheet_name_samples = book_name_samples.sheet_by_index(0)#根据顺序获取sheet这个是读取第一个sheet
    file = open('log.txt', 'w')
    #file1 = xlwt.Workbook()
    #excel_file=load_workbook(filename=r'E:\AI\ctc\log_template.xlsx')
    #ws = excel_file.get_sheet_by_name("Sheet2")
    row_excel = 1
    imgs=[]
    labels=[]    
    for row_name_samples in sheet_name_samples.get_rows():
        if str(row_name_samples[4].value) == mode_key:
            folder = str(row_name_samples[0].value)        
            answer = str(int(row_name_samples[2].value))                    
            all_path = images_dir + "/" + folder + '/' + str(row_name_samples[1].value)
            #os.path.exists
            if answer != -1 and os.path.isfile(all_path):
                #im = all_path
                #im_dapi = all_path_dapi
                #im_red = all_path_red
                print('reading the images:%s'%(all_path)," answer: ", str(answer))
                file.write('reading the images:%s'%(all_path)+" answer: "+ str(answer))
                #ws.cell(row=row_excel, column=1).value = folder
                #ws.cell(row=row_excel, column=2).value = str(answer)
                #row_excel += 1
                
                img=io.imread(all_path)
                img=transform.resize(img,(size_square,size_square))
                imgs.append(img)
                labels.append(answer)              
    file.close()            
    '''
    cate=[x for x in os.listdir(path) if os.path.isdir(path+x)]
    imgs=[]
    labels=[]
    for idx,folder in enumerate(cate):
        for im in glob.glob(path + folder+'/*.jpg'):
            print('reading the images:%s'%(im))
            img=io.imread(im)
            img=transform.resize(img,(w,h))
            imgs.append(img)

            labels.append(answer)
            
     '''       
    return np.asarray(imgs,np.float32),np.asarray(labels,np.int32)





def CTC_input_process(images_dir, labels_address, size_square, mode_key):
    

    data,label=read_img(images_dir, labels_address, size_square, mode_key)
    #打乱顺序
    num_example=data.shape[0]
    arr=np.arange(num_example)
    np.random.shuffle(arr)
    data=data[arr]
    label=label[arr]
    
    #将所有数据分为训练集和验证集
    ratio=1.0
    s=np.int(num_example*ratio)
    x_train=data[:s]
    y_train=label[:s]
    x_val=data[s:]
    y_val=label[s:]
    
    #print("x_train:",x_train,"y_train:",y_train)
    #ben
    #x_train=data
    #y_train=label
    #ben
    #train_model_v10(x_train,y_train)
    return  y_train, x_train

if __name__ == '__main__':

        print("###########################################this is beginning: \n")
        parser = argparse.ArgumentParser(description='manual to this script', epilog="authors of this script are PengChao YeZixuan XiaoYupei and Ben ")
        #parser.add_argument('--gpus', type=str, default = None)
        #parser.add_argument('--batch_size', type=int, default=32)
        parser.add_argument('--images_dir', type=str, default = "../../data/geneis/CTC/images")
        parser.add_argument('--labels_address', type=str, default = "../../data/geneis/CTC/result.xlsx")
        parser.add_argument('--save_model_address', type=str, default = "../../data/geneis/CTC/my_model.CNN_3")
        parser.add_argument('--size_square', type=int, default = 100)
        parser.add_argument('--label_types', type=str, default = 2)
        #parser.add_argument('--model_number', type=int, default=0,help="choose the model:0緾NN_3, 1縑GG_16, 2縭esnet_34,3縭esnet_50, 4縂oogleNet,5縄nception_V3, 6縄nception_V4,7縄nception_resnet_v1, 8縄nception_resnet_v2,9縎huffleNet")
        parser.add_argument('--model', type=str, default="CNN_3",help="choose the model:0緾NN_3, 1縑GG_16, 2縭esnet_34,3縭esnet_50, 4縂oogleNet,5縄nception_V3, 6縄nception_V4,7縄nception_resnet_v1, 8縄nception_resnet_v2,9縎huffleNet")
        parser.add_argument('--epochs', type=int, default=10)
        parser.add_argument('--times', type=int, default=2)
        parser.add_argument('--L1', type=int, default=4,help ="the number of the first conv2d")
        parser.add_argument('--L2', type=int, default=8,help = "the number of the second conv2d ")
        parser.add_argument('--F1', type=int, default=3,help = "the size of the first conv2d layer")
        parser.add_argument('--F2', type=int, default=2,help = "the size of the first maxpooling layer")
        parser.add_argument('--F3', type=int, default=3,help = "the size of the second conv2d layer")
        parser.add_argument('--n_splits', type=int, default=2, help="Cross validation")
        parser.add_argument('--cross_validation_method', type=int, default=0, help="1:KF.split 2:cross_val_predict 3:cross_val_score")
        parser.add_argument('--mode', type=str, default="train", help="train or test")

        parser.add_argument('--label_name', type=str, default="file_name,labels", help="label_name")
        parser.add_argument('--scan_window_suffix', type=str, default="*.png", help="scan_window_suffix")
        parser.add_argument('--image_num', type=int, default=1, help="The number of small images selected for each large image")
        parser.add_argument('--ID_prefix_num', type=int, default=15)
        parser.add_argument('--batch_size', type=int, default=2,help="batch_size")
        parser.add_argument('--tensorflow_version', type=str, default="2.0",help="batch_size")
        parser.add_argument('--roc_address', type=str, default="roc.png",help="roc_address")
        parser.add_argument('--roc_title', type=str, default="ROC curve",help="roc curve title")
        args = parser.parse_args()
        print("args.labels_address:",args.labels_address)
        print("args.images_dir:",args.images_dir)
        print("args.label_name:",args.label_name)
        model_types = ['CNN_3', 'VGG_16', 'resnet_34', 'resnet_50', 'GoogleNet', 'Inception_V3', 'Inception_V4',
                       'Inception_resnet_v1', 'Inception_resnet_v2','ShuffleNet']
        #save_model_address = args.save_model_address + model_types[args.model_number]
        save_model_address = args.save_model_address
        #label_list, image_data_list = CTC_input_process(args.images_dir, args.labels_address, args.size_square, args.label_name, args.image_num, args.ID_prefix_num, args.scan_window_suffix)

        #model_types[args.model_number]
        model_name = args.model
        print("----------------save_model_address:",save_model_address)
        if(args.tensorflow_version=="2.0"):
            label_list, image_data_list = CTC_input_process(args.images_dir,args.labels_address,args.size_square,mode_key=args.mode)
            #temp image_data_list  = image_data_list/255.0
            print("#############################################this is the end of read image! \n")
            print("image_data_list:",image_data_list.shape)
            image_data_list = np.reshape(image_data_list, (len(image_data_list), args.size_square, args.size_square, 3))
            if(args.mode=="train"):
                CNN.model_train(save_model_address, args.model, image_data_list, label_list, args.size_square, args.size_square, 3, args.label_types,args.epochs, args.times, args.L1,args.L2, args.F1, args.F2, args.F3,args.n_splits,args.cross_validation_method,args.batch_size,args.roc_address,args.roc_title)
            elif(args.mode=="test"):
                print("-------------------enter mode test---------------------")
                CNN.model_test(save_model_address, args.model, image_data_list, label_list,args.roc_address,args.roc_title)
        elif(args.tensorflow_version=="1.0"):
            #print(image_data_list)
            image_data_list = list(map(int, image_data_list))
            label_list = list(map(int, label_list))
            CNN_model_V1_0.train_model_v1_0(save_model_address, image_data_list, label_list)
 
        # draw_heatmap.draw_heatmap(args.images_dir, args.labels_address, args.need_save_WGI, args.size_square)


        ###########################################
        #tensor
        # data = np.reshape(image_data_list, (len(image_data_list), args.size_square, args.size_square, 4))
        #
        # y = tf.one_hot(label_list, depth=1)
        # ds = tf.data.Dataset.from_tensor_slices((data, y))
        # ds = ds.map(prepare_features_and_labels)
        # ds = ds.shuffle(10000).batch(10)
        #
        # sample = next(iter(ds))
        #
        # train_images= sample[0]
        # test_images = sample[0]
        # train_labels= sample[1]
        # test_labels = sample[1]
        # CNN.model_train(save_model_address, model_types[args.model_number], train_images, train_labels, test_images,
        #                 test_labels, args.size_square, args.size_square, 4, args.label_types, args.epochs, args.times,
        #                 args.L1, args.L2, args.F1, args.F2, args.F3, args.n_splits)

